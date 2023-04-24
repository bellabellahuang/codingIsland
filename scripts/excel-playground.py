import openpyxl as xl
import re
from copy import copy

FILTER_ROW_REGEX = '^Applied filters.*'
validator = re.compile(FILTER_ROW_REGEX)

MONTH_MAP = {
    'January':'01',
    'February':'02',
    'March':'03',
    'April':'04',
    'May':'05',
    'June':'06',
    'July':'07',
    'August':'08',
    'September':'09',
    'October':'10',
    'November':'11',
    'December':'12',
}

input ="original_file.xlsx"
wb = xl.load_workbook(input, data_only=True)
output = xl.Workbook()
tab_names = wb.sheetnames

for name in tab_names:
    if name in ['name_1', 'name_2']: continue # skip specific tabs
    tab = wb[name]
    print(f"Formatting {tab}")
    ws = output.create_sheet(title=name)

    # remove the "Applied filter" row
    for row in tab.iter_rows():
        cell_value = str(row[0].value)
        if validator.match(cell_value):
            tab.delete_rows(row[0].row, 1)
            print("Filter row deleted.")
            break

    # get all the column names
    headers = []
    for col in tab[1]:
        headers.append(col.value)

    # add the recalc difference col if it does not exist
    RECALC_DIFF_COL = "New col name"
    if len(headers) == 13 and RECALC_DIFF_COL not in headers:
        tab.insert_cols(11)
        print(F"{RECALC_DIFF_COL} col is inserted.")

    # for col in tab[1]:
    #     print(col.value)

    # add column for PID
    tab.insert_cols(3)
    for row in tab.iter_rows():
        row[2].value = name

    # save the new tab
    for row in tab.iter_rows():
        # if row[0].row == 4:
        for cell in row:
            # print(cell.row, row.index(cell))
            col = row.index(cell) + 1
            # keep header row and first 3 cols
            if cell.row ==1 or col in [1,2,3]:
                ws.cell(row=cell.row, column=col).value = MONTH_MAP.get(cell.value, cell.value)
            elif cell.has_style and cell.fill.bgColor.rgb != '00000000':
                # print(cell.fill.bgColor.rgb)
                ws.cell(row=cell.row, column=col).value = cell.value
                ws.cell(row=cell.row, column=col).fill = copy(cell.fill)

output.save(filename="new_file.xlsx")
print("File saved.")

# merge tabs
adj_wb = xl.load_workbook("new_file.xlsx", data_only=True)
single_wb = xl.Workbook()
single_ws = single_wb.create_sheet()

single_row = 0
adj_tabs = adj_wb.sheetnames
for name in adj_tabs:
    if name == 'Sheet': continue
    tab = adj_wb[name]
    for row in tab.iter_rows():
        for cell in row:
            col = row.index(cell) + 1
            single_ws.cell(row=single_row + cell.row, column=col).value = cell.value

    single_row += tab.max_row
    # single_col += tab.max_column

single_wb.save(filename="merge_file.xlsx")
print("Merged.")
